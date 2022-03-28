import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# Get the style from the other header column and copy to the newly added column
col_4 = product_list.cell(1, 4)
product_list.cell(1, 5).font = copy(col_4.font)
product_list.cell(1, 5).fill = copy(col_4.fill)
product_list.cell(1, 5).value = 'Inventory Price'


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_number)] = int(inventory)

    # Add a column to excel file for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# Haven't been able to get the column width thing to work yet...
# product_list.column_dimensions["E"].auto_size = True

inv_file.save("inventory_with_total_value.xlsx")
